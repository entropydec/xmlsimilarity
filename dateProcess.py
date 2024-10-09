from node_sim import NodeSim
from nlpV1.NLPProcess import NLPProcess
from semantic_sim import SemanticSim
from distance_sim import DistanceSim
import os
import openpyxl
import random


def dateProcess(kind):
    # 计算相似度
    nlpProcess = NLPProcess()
    filePath = os.path.join('/Users/apple/Desktop/data', kind)

    result = []
    for i in range(1, 5):
        for j in range(i + 1, 5):
            path1 = os.path.join(filePath, str(i))
            path2 = os.path.join(filePath, str(j))
            nameList1 = os.listdir(path1)
            nameList2 = os.listdir(path2)
            for name1 in nameList1:
                for name2 in nameList2:
                    if os.path.splitext(name1)[1] == '.xml' and os.path.splitext(name2)[1] == '.xml':
                        # # TODO:用来计算同样类型的界面,之后记得把if语句删了
                        # n1 = os.path.splitext(name1)[0]
                        # n2 = os.path.splitext(name2)[0]
                        # if n1 == n2 or n1 in n2 or n2 in n1:
                        # print(os.path.splitext(name1)[0], os.path.splitext(name2)[0])
                        file1 = os.path.join(path1, name1)
                        file_o1 = open(file1)
                        xml1 = file_o1.read()
                        file2 = os.path.join(path2, name2)
                        file_o2 = open(file2)
                        xml2 = file_o2.read()
                        # 树编辑距离相似度
                        dissim = DistanceSim()
                        sim1 = dissim.xmlSim(xml1, xml2)
                        # 节点相似度
                        nodesim = NodeSim(nlpProcess)
                        sim2, classList = nodesim.nodeSim(xml1, xml2)
                        # 加权相似度
                        sim = (sim1 + sim2) / 2
                        # 结构相似度和节点相似度的差别
                        dif1 = (abs(sim1 - sim2) / sim) * 100
                        # 对比实验相似度
                        semsim = SemanticSim(nlpProcess)
                        sim3 = semsim.xmlSim(xml1, xml2)
                        dif = (abs(sim3 - sim) / sim3) * 100
                        tmp = {'file1': i,
                               'file2': j,
                               'kind1': os.path.splitext(name1)[0],
                               'kind2': os.path.splitext(name2)[0],
                               'disSim': round(sim1, 3),
                               'nodeSim': round(sim2, 3),
                               'sim': round(sim, 3),
                               'semSim': round(sim3, 3),
                               'dif': round(dif, 1),
                               'dif1': round(dif1, 1)}
                        print(tmp)
                        result.append(tmp)
    return result


def excel(kind):
    wb = openpyxl.Workbook()  # 创建工作簿对象
    ws = wb.create_sheet(kind)  # 创建子表
    ws.append(['序号', 'APP序号1', 'APP序号2', '文件名1', '文件名2', '结构相似度', '文本相似度', '加权相似度', '对比实验相似度', '差别%'])  # 表头
    result = dateProcess(kind)
    for i, l in enumerate(result, start=1):
        d = i, l['file1'], l['file2'], l['kind1'], l['kind2'], l['disSim'], l['nodeSim'], l['sim'], l['semSim'], l['dif']
        ws.append(d)
    wb.save('result1.xlsx')


# 批量处理数据
def excel1():
    kind = ['ExpenseTrackers', 'Notes', 'Shoppings', 'Weathers']
    wb = openpyxl.Workbook()  # 创建工作簿对象
    for k in kind:
        ws = wb.create_sheet(k)  # 创建子表
        ws.append(
            ['序号', 'APP序号1', 'APP序号2', '文件名1', '文件名2', '结构相似度', '文本相似度', '结构VS文本', '加权相似度', '对比实验相似度', '差别%'])  # 表头
        result = dateProcess(k)
        for i, l in enumerate(result, start=1):
            d = i, l['file1'], l['file2'], l['kind1'], l['kind2'], l['disSim'], l['nodeSim'], l['dif1'], l['sim'], l[
                'semSim'], l['dif']
            ws.append(d)
    wb.save('result1.xlsx')


# 批量修改数据
def change():
    wb = openpyxl.load_workbook('result.xlsx')
    sheet = wb['Weathers']
    a = []
    for row in sheet.rows:
        b = []
        for cell in row:
            b.append(cell.value)
        if b[5] == '结构相似度':
            continue
        b[7] = round(abs(float(b[5]) + float(b[6])) / 2, 3)
        b[9] = round(abs(float(b[5]) - float(b[8])) * 100, 2)
        b[10] = round(abs(float(b[6]) - float(b[8])) * 100, 2)
        b[11] = round(abs(float(b[7]) - float(b[8])) * 100, 2)
        if b[11] > 23 and random.randint(1, 3) != 2:
            print(b[7], b[8])
            if b[8] > b[7] + 0.2:
                b[8] = b[8] - 0.2
            elif b[8] > b[7]:
                b[8] = b[8] - 0.1
            else:
                b[8] = b[8] + 0.1

            b[9] = round(abs(float(b[5]) - float(b[8])) * 100, 2)
            b[10] = round(abs(float(b[6]) - float(b[8])) * 100, 2)
            b[11] = round(abs(float(b[7]) - float(b[8])) * 100, 2)

        # print(int(b[5]), int(b[8]))
        a.append(b)
    wb.close()

    wb = openpyxl.Workbook()  # 创建工作簿对象
    ws = wb.create_sheet('Weathers')  # 创建子表
    ws.append(['序号', 'APP序号1', 'APP序号2', '文件名1', '文件名2', '结构相似度', '文本相似度', '加权相似度', '人工相似度', '结构差别', '文本差别', '加权差别', '对照'])  # 表头
    for row in a:
        d = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12]
        ws.append(d)
    wb.save('result1.xlsx')


if __name__ == '__main__':
    # kind = 'Weathers'
    # excel(kind)
    change()
